#!/home/joonho/anaconda3/bin/python

import argparse
import re
import os
from common import f_ext
from openpyxl import load_workbook
from wcwidth import wcswidth

def pad_right_display(s, width):
    s = str(s).strip()
    display = ''
    current_width = 0
    for ch in s:
        ch_width = wcswidth(ch)
        if current_width + ch_width > width:
            break
        display += ch
        current_width += ch_width
    padding = width - wcswidth(display)
    return display + ' ' * max(padding, 0)

def extract_excel(xlsx, ext):

    # Load the workbook and sheet
    wb = load_workbook(xlsx)
    sheet = wb.active

    # Read header (assumes 1st row is header)
    original_headers = [cell.value.strip() if isinstance(cell.value, str) else str(cell.value) for cell in sheet[1]]
    headers = ['Full Name'] + original_headers[2:]  # Combine first two columns

    # Read data and build new rows with "Full Name"
    data_rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        full_name = f"{row[0]} {row[1]}".strip()
        new_row = (full_name,) + row[2:]
        data_rows.append(new_row)

    # Set max width
    MAX_WIDTH = 25

    # Determine max length per column, capped at MAX_WIDTH
    num_cols = len(headers)
    max_lengths = [min(wcswidth(str(headers[i])), MAX_WIDTH) for i in range(num_cols)]

    for row in data_rows:
        for i in range(num_cols):
            cell = str(row[i]).strip() if row[i] is not None else ""
            cell_width = wcswidth(cell)
            max_lengths[i] = min(max(max_lengths[i], cell_width), MAX_WIDTH)


    print(f"max length {max_lengths}") 
    # Format row with stripping and right-alignment
    def format_row(row, max_lengths):
        return ' '.join([pad_right_display(cell, max_lengths[i]) for i, cell in enumerate(row)])
    # Print header
    print(format_row(headers, max_lengths))
    print("-" * (sum(max_lengths) + 3 * (num_cols - 1)))

    # Print data rows
    for row in data_rows:
        print(format_row(row, max_lengths))

    return 0

def main():
    parser = argparse.ArgumentParser(description='To get rid of abnormal DOS at start energy')
    parser.add_argument('file', nargs='?', default='DOSCAR', help='read DOSCAR')
    parser.add_argument('-o', '--old', default='_o', help='save original DOSCAR to DOSCAR_o')
    args = parser.parse_args()

    if f_ext(args.file) != 'xlsx':
        print(f"Warning: file is not extension of xlsx")
    
    extract_excel(args.file, args.old) 

if __name__ == '__main__':
    main()
