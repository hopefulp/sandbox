#!/home/joonho/anaconda3/bin/python

import argparse
import re
import os
from common import f_ext
from wcwidth import wcswidth

def pad_right_display(s, width):
    s = str(s).strip()
    display = ''
    current_width = 0
    for ch in s:
        ch_width = wcswidth(ch)
        if current_width + ch_width > width:
            break
        display += ch
        current_width += ch_width
    padding = width - wcswidth(display)
    return display + ' ' * max(padding, 0)

def deduplicate_keep_offline_first(rows):
    unique_entries = {}   # key: email, value: row
    removed = []

    for row in rows:
        email = str(row[1]).strip().lower()
        status = str(row[3]).strip().lower() if row[3] else ''

        if email not in unique_entries:
            unique_entries[email] = row  # Add first time
        else:
            existing_status = str(unique_entries[email][3]).strip().lower()
            if existing_status == 'online' and status == 'offline':
                removed.append(unique_entries[email])     # Remove the online one
                unique_entries[email] = row               # Replace with offline
            else:
                removed.append(row)  # Remove the duplicate (same or lower priority)

    return list(unique_entries.values()), removed



def extract_table_file(fname, isort, Loption):

    if f_ext(fname) == 'xlsx':
        from openpyxl import load_workbook

        # Load the workbook and sheet
        wb = load_workbook(fname)
        sheet = wb.active

        # Read header (assumes 1st row is header)
        original_headers = [cell.value.strip() if isinstance(cell.value, str) else str(cell.value) for cell in sheet[1]]
        headers = ['Full Name'] + original_headers[2:]  # Combine first two columns

        # Read data and build new rows with "Full Name"
        data_rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            full_name = f"{row[0]} {row[1]}".strip()
            new_row = (full_name,) + row[2:]
            data_rows.append(new_row)
    elif f_ext(fname) == 'csv':
        import csv
        with open(fname, encoding='utf-8') as f:
            reader = csv.reader(f)
            rows = list(reader)

        # Get headers
        original_headers = rows[0]
        headers = ['Full Name'] + original_headers[2:]

        # Build data rows
        data_rows = []
        for row in rows[1:]:
            full_name = f"{row[0]} {row[1]}".strip()
            new_row = (full_name,) + tuple(row[2:])
            data_rows.append(new_row)


    # Set max width
    MAX_WIDTH = 25

    # Determine max length per column, capped at MAX_WIDTH
    num_cols = len(headers)
    max_lengths = [min(wcswidth(str(headers[i])), MAX_WIDTH) for i in range(num_cols)]

    for row in data_rows:
        for i in range(num_cols):
            cell = str(row[i]).strip() if row[i] is not None else ""
            cell_width = wcswidth(cell)
            max_lengths[i] = min(max(max_lengths[i], cell_width), MAX_WIDTH)

    print(f"max length {max_lengths}") 
    # Format row with stripping and right-alignment
    def format_row(row, max_lengths, index=None):
        formatted = []
        if index is not None:
            index_str = f"{index:>3}  "  # Right-align index in 3 spaces
            formatted.append(index_str)
        for i, cell in enumerate(row):
            formatted.append(pad_right_display(cell, max_lengths[i]))
        return " ".join(formatted)

        #return ' '.join([pad_right_display(cell, max_lengths[i]) for i, cell in enumerate(row)])
    ### filter
    # Remove rows where the 4th column (index 3) is None
    data_rows = [row for row in data_rows if row[3] is not None]

    # Print header
    print("    " + format_row(headers, max_lengths))
    print("-" * (sum(max_lengths) + 3 * (num_cols - 1) + 5))

    ### Unique row: If duplicate id/email though name is different
    filtered_rows, removed_rows = deduplicate_keep_offline_first(data_rows)

    data_rows = filtered_rows
    # Print data rows
    sorted_rows = sorted(data_rows, key=lambda x: str(x[3]).strip().lower())
    for idx, row in enumerate(sorted_rows):
        if Loption:
            if row[3] == 'offline':
                print(format_row(row, max_lengths, index=idx+1))
        else:
             print(format_row(row, max_lengths, index=idx+1))

    return 0

def main():
    parser = argparse.ArgumentParser(description='arragne excel file')
    parser.add_argument('file',  help='input excel with extension of xlsx')
    parser.add_argument('-s', '--isort', help='sort by index')
    parser.add_argument('-o', '--option', action='store_true', help='if option, print only "offline"')
    args = parser.parse_args()

    if f_ext(args.file) not in ('xlsx', 'csv'):
        print(f"Warning: file is not an Excel or CSV file")
    
    extract_table_file(args.file, args.isort, args.option) 

if __name__ == '__main__':
    main()
